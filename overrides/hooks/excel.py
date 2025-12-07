import os
from shutil import copyfile
from textwrap import dedent

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation

TAGS = [
    "event", "messe", "museum", "park", "theater"
]
FILES = [
    "webshop.yml",
    "produkte.yml",
    "abo.yml",
    "b2b.yml",
    "aussteller.yml",
    "rechnungswesen.yml",
    "backend.yml",
    "ressourcen.yml",
    "kasse.yml",
    "zutrittskontrolle.yml",
    "automat.yml",
    "reporting.yml",
    "schnittstellen.yml",
    "saas.yml",
    "onpremise.yml",
    "support.yml",
]
LVL = {
    "MUST": "MUSS",
    "SHOULD": "SOLL",
    "MAY": "KANN",
}

def _criteria(ws, cat, data, tags, level, parent=""):
    i = ws.max_row + 1
    ws.append([
        "=ROW()-1",
        cat,
        parent + data["title"],
        data["description"].strip(),
        LVL[data["level"].upper()],
        "",
        f'=IF(E{i}="MUSS", "", IF(E{i}="SOLL", IF(F{i}="Im Standard erfüllt", 5, IF(F{i}="Mit Zusatzkosten erfüllt", 4, 0)), IF(E{i}="KANN", IF(F{i}="Im Standard erfüllt", 2, IF(F{i}="Mit Zusatzkosten erfüllt", 1, 0)), "")))',
        "",
        "",
    ])
    if data.get("comment"):
        ws["D"][ws.max_row - 1].comment = Comment(
            text=data["comment"].strip(),
            author="ticketing-ausschreibung.de",
            width=400,
            height=100,
        )
    fontStyle = Font(name="Calibri", sz=10)
    for c in "ABCDEFGHI":
        ws[c][ws.max_row - 1].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left" if c not in "GH" else "right")
        if c in "FHI":
            ws[c][ws.max_row - 1].fill = PatternFill(start_color='FFDEDEFF', end_color='FFDEDEFF', fill_type='solid')
        ws[c][ws.max_row - 1].font = fontStyle
    ws["H"][ws.max_row - 1].number_format = '#,##0.00€'

    if data.get("sub"):
        for l in data["sub"]:
            if l.get("tags"):
                if not any(t in l["tags"] for t in tags):
                    continue
            _criteria(ws, cat, l, tags, level + 1, parent + data["title"] + " – ")


def _chapter(ws, title, requirements, tags):
    # if not any(any(not l.get("tags") or t in l["tags"] for t in tags) for l in requirements):
    #     return
    # ws.append(['=SUMPRODUCT(--(K:K="A")*--(ROW(K:K)<=ROW()))', title, title, "", "", "", "", "", "A"])

    for l in requirements:
        if l.get("tags"):
            if not any(t in l["tags"] for t in tags):
                continue
        _criteria(ws, title, l, tags, 1)


def build_price_sheet(wb, tag):
    ws = wb.create_sheet("Preisblatt")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.freeze_panes = "A2"
    main_sheet = wb["Anforderungen"]

    ws.append([
        "Lfd. Nr.",  # A
        "Kategorie",  # B
        "Position",  # C
        "Menge",  # D
        "Einheit",  # E
        "Einzelpreis",  # F
        "Summe (netto)",  # G
    ])
    for c in range(1, 10):
        fontStyle = Font(bold=True, name="Calibri")
        ws.cell(row=1, column=c).font = fontStyle

    rows = [
        ["Software", "Einmalige Lizenzkosten", "1", "pauschal", ""],
        ["Software", "Jährliche Lizenzkosten", "5", "Jahre", ""],
        ["Software", "Jährliche Hosting-Kosten (bei SaaS)", "5", "Jahre", ""],
        ["Software", "Kosten pro Ticketverkauf", "50000", "Tickets", ""],
        ["Software", "Transaktionsentgelt pro Ticketverkauf", "1000000", "% auf €", ""],
        ["Implementierung", "Einführungskosten", "1", "pauschal", ""],
        ["Implementierung", "Basissetup", "1", "pauschal", ""],
        ["Implementierung", "Zusatzaufwände aus Anforderungen", "1", "pauschal", f"=Anforderungen!H{main_sheet.max_row}"],
        ["Support", "Kosten für sonstige Implementierungsleistungen auf Abruf", "32", "Stunden", ""],
        ["Support", "Kosten für initiale Schulung der Key-User", "1", "pauschal", ""],
        ["Support", "Kosten für weitere Schulungsleistungen auf Abruf", "32", "Stunden", ""],
        ["Support", "Kosten für sonstige Beratungsleistungen auf Abruf", "48", "Stunden", ""],
        ["Support", "Monatliche Kosten für Support/Hotline", "60", "Monate", ""],
        ["Hardware", "Mobiles Kassensystem lt. Anforderungskatalog", "3", "Stück", ""],
        ["Hardware", "Stationäres Kassensystem lt. Anforderungskatalog", "5", "Stück", ""],
        ["Hardware", "Belegdrucker lt. Anforderungskatalog", "5", "Stück", ""],
        ["Hardware", "Ticketdrucker lt. Anforderungskatalog", "5", "Stück", ""],
        *([["Hardware", "Plastikkartendrucker lt. Anforderungskatalog", "5", "Stück", ""]] if tag in ("museum", "park", "event") else []),
        ["Hardware", "Handscanner lt. Anforderungskatalog", "10", "Stück", ""],
        *([["Hardware", "Self-Badging-Station inkl. Badgedrucker lt. Anforderungskatalog", "5", "Stück", ""]] if tag in ("messe",) else []),
        *([["Hardware", "Badgedrucker lt. Anforderungskatalog", "10", "Stück", ""]] if tag in ("messe",) else []),
        *([["Hardware", "Stationäres Drehkreuz lt. Anforderungskatalog", "2", "Stück", ""]] if tag in ("messe", "museum", "park", "event") else []),
        *([["Hardware", "Stationäres Drehkreuz inkl. Badgedrucker lt. Anforderungskatalog", "15", "Stück", ""]] if tag in ("messe",) else []),
        *([["Hardware", "Mobiles Drehkreuz lt. Anforderungskatalog", "10", "Stück", ""]] if tag in ("messe",) else []),
        *([["Hardware", "Ticketautomat outdoor cashless", "1", "Stück", ""]] if tag in ("park",) else []),
        *([["Hardware", "Ticketautomat outdoor mit Bargeld", "1", "Stück", ""]] if tag in ("park",) else []),
        *([["Hardware", "Ticketautomat indoor cashless", "1", "Stück", ""]] if tag in ("museum", "park") else []),
        *([["Hardware", "Ticketautomat indoor mit Bargeld", "1", "Stück", ""]] if tag in ("park",) else []),
    ]

    for r in rows:
        ws.append(["", *r, f"=D{ws.max_row + 1}*F{ws.max_row + 1}"])
        for c in "ABCDEFGH":
            if c in "ABC":
                ws[c][ws._current_row - 1].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            elif c in "DE":
                ws[c][ws._current_row - 1].alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            elif c in "FG":
                ws[c][ws._current_row - 1].alignment = Alignment(wrap_text=True, vertical="top", horizontal="right")
                if c == "F" and "%" in r[3]:
                    ws[c][ws._current_row - 1].number_format = '#,##0.00%'
                else:
                    ws[c][ws._current_row - 1].number_format = '#,##0.00€'
            if c in "F":
                ws[c][ws._current_row - 1].fill = PatternFill(start_color='FFDEDEFF', end_color='FFDEDEFF', fill_type='solid')

    ws.append([
        "",
        "Summe",
        "",
        "",
        "",
        "",
        f'=SUM(G2:G{ws.max_row})',
    ])
    for c in range(1, 10):
        fontStyle = Font(bold=True, name="Calibri")
        ws.cell(row=ws.max_row, column=c).font = fontStyle
    ws.cell(row=ws.max_row, column=7).number_format = '#,##0.00€'


def build_xlsx(tag):
    wb = Workbook()
    ws = wb.active
    ws.title = "Anforderungen"

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 30
    ws.freeze_panes = "A2"
    ws.append([
        "Lfd. Nr.",  # A
        "Kategorie",  # B
        "Anforderung",  # C
        "Beschreibung",  # D
        "Art",  # E
        "Erfüllungsgrad",  # F
        "Bewertung",  # G
        "Zusatzkosten",  # H
        "Kommentar Anbieter",  # I
    ])

    for c in range(1, 10):
        fontStyle = Font(bold=True, name="Calibri", sz=10)
        ws.cell(row=1, column=c).font = fontStyle

    for file in FILES:
        with open("data/modules/" + file, "r") as f:
            data = yaml.safe_load(f.read())
        _chapter(ws, data["title"], data["requirements"], [tag])

    dv = DataValidation(type="list", formula1='"MUSS,SOLL,KANN"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('E1:E1048576')

    dv = DataValidation(type="list", formula1='"Im Standard erfüllt,Mit Zusatzkosten erfüllt,Nicht erfüllt"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('F1:F1048576')

    rule = CellIsRule(operator='equal', fill=PatternFill(start_color='FFFFC1DF', end_color='FFFFC1DF', fill_type='solid'), formula=['"MUSS"'], stopIfTrue=True)
    ws.conditional_formatting.add('E1:E1048576', rule)
    rule = CellIsRule(operator='equal', fill=PatternFill(start_color='FFFFF2B8', end_color='FFFFF2B8', fill_type='solid'), formula=['"SOLL"'], stopIfTrue=True)
    ws.conditional_formatting.add('E1:E1048576', rule)
    rule = CellIsRule(operator='equal', fill=PatternFill(start_color='FFA7DAFF', end_color='FFA7DAFF', fill_type='solid'), formula=['"KANN"'], stopIfTrue=True)
    ws.conditional_formatting.add('E1:E1048576', rule)

    rule = CellIsRule(operator='equal', fill=PatternFill(start_color='FFFFC1DF', end_color='FFFFC1DF', fill_type='solid'), formula=['"Nicht erfüllt"'], stopIfTrue=True)
    ws.conditional_formatting.add('F1:F1048576', rule)
    rule = CellIsRule(operator='equal', fill=PatternFill(start_color='FFFFF2B8', end_color='FFFFF2B8', fill_type='solid'), formula=['"Mit Zusatzkosten erfüllt"'], stopIfTrue=True)
    ws.conditional_formatting.add('F1:F1048576', rule)
    rule = CellIsRule(operator='equal', fill=PatternFill(start_color='FFCCFFB9', end_color='FFCCFFB9', fill_type='solid'), formula=['"Im Standard erfüllt"'], stopIfTrue=True)
    ws.conditional_formatting.add('F1:F1048576', rule)

    ws.append([
        "",
        "Summe",
        "Summe",
        "",
        "",
        "",
        f'=SUM(G2:G{ws.max_row})',
        f'=SUM(H2:H{ws.max_row})',
        "",
    ])
    for c in range(1, 10):
        fontStyle = Font(bold=True, name="Calibri", sz=10)
        ws.cell(row=ws.max_row, column=c).font = fontStyle
    ws.cell(row=ws.max_row, column=8).number_format = '#,##0.00€'

    build_price_sheet(wb, tag)

    ws = wb.create_sheet("Einleitung", index=0)
    ws["A1"].value = dedent("""
    Dies ist eine Ausschreibungs-Vorlage für Ticketing-Ausschreibungen für öffentliche Einrichtungen, die auf www.ticketing-ausschreibung.de herungergeladen wurde.
    
    Die Inhalte stehen unter der Lizenz CC0 1.0 Universal, was bedeutet, dass sie komplett frei verwendet werden dürfen, auch ohne Nennung der Urheber.
    
    Wir empfehlen, dass Sie sich ausführlich mit der Thematik beschäftigen und die Anforderungen präzisieren, eigene Anforderungen ergänzen und Anforderungen streichen.
    
    Auch das Berechnungsschema der Bewertung, sowie das Preisblatt müssen zwingend von Ihnen auf Ihre konkreten Bedürfnisse angepasst werden.
    
    Beachten Sie dabei auch unsere Tipps für eine gute Ausschreibung auf www.ticketing-ausschreibung.de.
    
    Die in der Excel-Datei hellblau hinterlegten Zellen sind zur Ausfüllung durch die Bieter gedacht. Wir empfehlen, nach Abschluss der Bearbeitung alle anderen Felder für Bearbeitung durch die Bieter zu sperren.
    
    Wir übernehmen keine Haftung für die Verwendung dieser kostenlos bereitgestellten Vorlage. Die Zurverfügungstellung dieser Vorlage stellt keine (Rechts)beratung dar.
    """).strip()
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.column_dimensions["A"].width = 90

    return wb

def on_files(files, config, *args, **kwargs):
    # Ensure presence of cache directory
    cache = ".cache/plugin/excel"
    if not os.path.isdir(cache):
        os.makedirs(cache)
    dest_dir = os.path.join(
        config.site_dir,
        "download",
    )
    if not os.path.isdir(dest_dir):
        os.makedirs(dest_dir)

    for t in TAGS:
        cache_path = os.path.join(cache, f"cache.xlsx")
        dest_path = "{}.xlsx".format(os.path.join(
            dest_dir,
            f"ausschreibung-{t}"
        ))
        print("Building xlsx for", t)
        wb = build_xlsx(t)
        wb.save(cache_path)
        copyfile(cache_path, dest_path)
        print("Built xlsx for", t)
